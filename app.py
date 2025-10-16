
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime

# Constants for styling
header_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Function to apply header layout
def apply_header(ws):
    ws.merge_cells('A1:I1')
    ws['A1'] = "DYNATRADE AUTOMOTIVE LLC - QUOTATION"
    ws.merge_cells('A2:B2')
    ws['A2'] = "Customer Code"
    ws.merge_cells('A3:B3')
    ws['A3'] = "Customer Name"
    ws.merge_cells('C2:E2')
    ws['C2'] = ""
    ws.merge_cells('C3:E3')
    ws['C3'] = ""
    ws.merge_cells('F2:G3')
    ws['F2'] = "Date:"
    ws.merge_cells('H2:I3')
    ws['H2'] = datetime.datetime.today().strftime("%d/%m/%Y")

    # Apply styles to header cells
    for row in [1, 2, 3, 4]:
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

    # Set column headers in row 4
    headers = ["S.No", "Inquired Part No", "Part Number", "Manf.Part", "Description", "Brand", "Stock on Hand", "Unit Price", "COO"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header

# Function to write data to worksheet
def write_data(ws, data):
    current_row = 5
    previous_serial = None
    for row in data:
        serial = row[0]
        if previous_serial is not None and serial != previous_serial:
            for col in range(1, 10):
                blank_cell = ws.cell(row=current_row, column=col)
                blank_cell.border = thin_border
            current_row += 1
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.border = thin_border
        previous_serial = serial
        current_row += 1

# Function to adjust column widths
def adjust_column_widths(ws):
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

# Streamlit app
st.title("Dynatrade Quotation Formatter")
uploaded_file = st.file_uploader("Upload Excel Dump File", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file, engine='openpyxl')
    required_columns = ["S.No", "Inquired Part No", "Part Number", "Manf.Part", "Description", "Brand", "Stock on Hand", "Unit Price", "COO"]
    df_filtered = df[required_columns]

    # Create workbook and sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Quotation"
    apply_header(ws1)
    quotation_data = df_filtered[(df_filtered["Stock on Hand"] > 0) & (df_filtered["Brand"].notna())].values.tolist()
    write_data(ws1, quotation_data)
    adjust_column_widths(ws1)

    # Zero Stock Sheet
    ws2 = wb.create_sheet(title="Zero Stock")
    apply_header(ws2)
    zero_stock_data = df_filtered[(df_filtered["Stock on Hand"] == 0) & (df_filtered["Brand"].notna())]
    zero_stock_data = zero_stock_data[~zero_stock_data["S.No"].isin(df_filtered[df_filtered["Stock on Hand"] > 0]["S.No"])]
    write_data(ws2, zero_stock_data.values.tolist())
    adjust_column_widths(ws2)

    # Yellow Card Sheet
    ws3 = wb.create_sheet(title="Yellow Card")
    apply_header(ws3)
    yellow_card_data = df_filtered[df_filtered["Brand"].isna()]
    write_data(ws3, yellow_card_data.values.tolist())
    adjust_column_widths(ws3)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        label="Download Formatted Quotation File",
        data=output,
        file_name="Standardized_Quotation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
