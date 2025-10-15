import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

def apply_formatting(ws):
    yellow_fill = PatternFill(start_color="FFF6D6", end_color="FFF6D6", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header rows
    ws.merge_cells("A1:I1")
    ws["A1"] = "DYNATRADE AUTOMOTIVE LLC - QUOTATION"
    ws["A1"].fill = yellow_fill
    ws["A1"].font = bold_font
    ws["A1"].alignment = center_alignment

    ws.merge_cells("A2:B2")
    ws["A2"] = "Customer Code"
    ws["A2"].fill = yellow_fill
    ws["A2"].font = bold_font
    ws["A2"].alignment = center_alignment

    ws.merge_cells("C2:E2")
    ws["C2"].fill = yellow_fill
    ws["C2"].font = bold_font
    ws["C2"].alignment = center_alignment

    ws.merge_cells("A3:B3")
    ws["A3"] = "Customer Name"
    ws["A3"].fill = yellow_fill
    ws["A3"].font = bold_font
    ws["A3"].alignment = center_alignment

    ws.merge_cells("C3:E3")
    ws["C3"].fill = yellow_fill
    ws["C3"].font = bold_font
    ws["C3"].alignment = center_alignment

    ws.merge_cells("F2:G3")
    ws["F2"] = "Date:"
    ws["F2"].fill = yellow_fill
    ws["F2"].font = bold_font
    ws["F2"].alignment = center_alignment

    ws.merge_cells("H2:I3")
    ws["H2"] = datetime.today().strftime("%d/%m/%Y")
    ws["H2"].fill = yellow_fill
    ws["H2"].font = bold_font
    ws["H2"].alignment = center_alignment

    headers = ["S.No", "Inquired Part No", "Part Number", "Manf.Part", "Description", "Brand", "Stock on Hand", "Unit Price", "COO"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_alignment

def insert_data(ws, df):
    current_row = 5
    previous_serial = None
    for _, row in df.iterrows():
        serial = row["S.No"]
        if previous_serial is not None and serial != previous_serial:
            current_row += 1  # empty row between groups
        for col_num, value in enumerate([
            row["S.No"], row["Inquired Part No"], row["Part Number"], row["Manf.Part"],
            row["Description"], row["Brand"], row["Stock on Hand"], row["Unit Price"], row["COO"]
        ], 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
        previous_serial = serial
        current_row += 1

def process_file(uploaded_file):
    df = pd.read_excel(uploaded_file, engine="openpyxl")
    quotation_df = df.copy()
    zero_stock_df = quotation_df[quotation_df["Stock on Hand"] == 0].copy()
    yellow_card_df = quotation_df[quotation_df["Brand"].isna()].copy()
    quotation_df = quotation_df[~quotation_df.index.isin(zero_stock_df.index)]

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Quotation"
    apply_formatting(ws1)
    insert_data(ws1, quotation_df)

    ws2 = wb.create_sheet("Zero Stock")
    apply_formatting(ws2)
    insert_data(ws2, zero_stock_df)

    ws3 = wb.create_sheet("Yellow Card")
    apply_formatting(ws3)
    insert_data(ws3, yellow_card_df)

    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

st.title("Quotation Formatter App")
uploaded_file = st.file_uploader("Upload Excel Dump File", type=["xlsx"])
if uploaded_file:
    formatted_data = process_file(uploaded_file)
    st.download_button("Download Formatted Quotation", data=formatted_data, file_name="Formatted_Quotation.xlsx")
